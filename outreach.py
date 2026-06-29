"""
Trees Engineering — Outreach Automation
Usage:
  python outreach.py --mode drafts   # Cree les brouillons Gmail (defaut)
  python outreach.py --mode send     # Envoie directement (45s entre chaque)
  python outreach.py --mode schedule # Installe la tache planifiee Windows
"""
import argparse, sys, os, re, base64, imaplib, smtplib, time
import email as email_lib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from datetime import date, datetime, timezone
from dotenv import load_dotenv
import pandas as pd
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
load_dotenv(os.path.join(BASE_DIR, ".env"))

GMAIL        = os.getenv("GMAIL_ADDRESS")
APP_PASSWORD = os.getenv("GMAIL_APP_PASSWORD")
EXCEL_PATH   = os.getenv("EXCEL_PATH", "")
TODAY        = date.today()

# ── SECTOR & PROFILE HELPERS ──────────────────────────────────────────────────

def get_version(title):
    if not isinstance(title, str): return "C"
    t = title.lower()
    if any(k in t for k in ["hr","human resources","people","talent","recruitment"]): return "A"
    if any(k in t for k in ["engineer","technical","project","operations","epc",
                              "process","software","digital","research","well"]): return "B"
    return "C"

def get_sector(company):
    if not isinstance(company, str): return "the engineering sector"
    c = company.upper()
    if any(k in c for k in ["PETRONAS","SHELL","EDRA","SAPURA","MISC"]): return "Oil & Gas"
    if any(k in c for k in ["YTL","SUNWAY","BOUSTEAD","GAMUDA","IJM"]): return "Infrastructure & Construction"
    if any(k in c for k in ["TNB","TENAGA"]): return "Energy"
    if any(k in c for k in ["IHH","HOSPITAL","HEALTH"]): return "Healthcare"
    return "the engineering sector"

def parse_date(status):
    if not isinstance(status, str): return None
    m = re.search(r"\((\d{1,2}/\d{1,2})\)", status)
    if not m: return None
    try:
        day, month = m.group(1).split("/")
        return date(TODAY.year, int(month), int(day))
    except: return None

def get_sequence(days):
    if 5  <= days <= 7:  return "Follow-up 1"
    if 12 <= days <= 14: return "Follow-up 2"
    if 20 <= days <= 25: return "Final Email"
    return None

def first_name(full): return (full or "").strip().split()[0] if full else "there"

# ── TEMPLATE BUILDER ──────────────────────────────────────────────────────────

SUBJECTS = {
    ("Follow-up 1", "A"): "Quick question, {first_name}",
    ("Follow-up 1", "B"): "{first_name}, worth a quick chat?",
    ("Follow-up 1", "C"): "Checking in, {first_name}",
    ("Follow-up 2", "A"): "Finding engineers in {sector} right now",
    ("Follow-up 2", "B"): "One thing I keep hearing from {sector} teams",
    ("Follow-up 2", "C"): "Something relevant for {company}",
    ("Final Email",  "A"): "Closing the loop",
    ("Final Email",  "B"): "Closing the loop",
    ("Final Email",  "C"): "Closing the loop",
}

def get_subject(sequence, version, first_name_str, company, sector):
    tpl = SUBJECTS.get((sequence, version), "Trees Engineering")
    return tpl.format(first_name=first_name_str, company=company, sector=sector)

def p(text, bottom=14):
    return f'<p style="margin:0 0 {bottom}px;font-size:14px;color:#444;line-height:1.7;">{text}</p>'

def bold(text):
    return f'<strong style="color:#0D1F35;">{text}</strong>'

def corps_block(content):
    return f'<tr><td style="padding:36px 40px 0;">{content}</td></tr>'

FOLLOWUP1_CORPS = {
"A": corps_block(
    p("Hi {first_name},") +
    p("I reached out last week about Trees Engineering and wanted to circle back.") +
    p("Sourcing the right engineering talent for niche roles is one of the hardest things HR teams deal with in {sector} right now. Most agencies send CVs that don't fit. The process drags on.") +
    p("We built Trees specifically to fix that. Pre-vetted engineers, available fast, for exactly the role you need.", bottom=0)
),
"B": corps_block(
    p("Hi {first_name},") +
    p("Following up on my email from last week.") +
    p("I know how it goes on project teams: scope shifts, a specialist is missing for the FEED phase, and the usual hiring process takes weeks you don't have.") +
    p("Trees Engineering exists for those exact moments. We place specialist engineers fast, without the agency back-and-forth.", bottom=0)
),
"C": corps_block(
    p("Hi {first_name},") +
    p("Just checking this didn't get lost in your inbox.") +
    p("Trees Engineering works with firms like Technip, Worley and Subsea7 on engineering talent. The pitch is simple: right engineer, fast, no agency markup.") +
    p("Would 20 minutes with Alex, our COO, be worth it?", bottom=0)
),
}

FOLLOWUP2_CORPS = {
"A": corps_block(
    p("Hi {first_name},") +
    p("I keep hearing the same thing from HR leaders in {sector}: they're not short of CVs, they're short of people who actually know what they're doing on site.") +
    p("That's the problem Trees Engineering was built around. We don't source generalists. Every engineer on our platform comes with verified project experience in your sector.") +
    p("Happy to show you how it works in 20 minutes. Would that be useful?", bottom=0)
),
"B": corps_block(
    p("Hi {first_name},") +
    p("One thing I keep hearing from project managers in {sector}: flexible engineering talent sounds great in theory, but finding someone who can hit the ground running is another story.") +
    p("That's where Trees is different. Our engineers are pre-vetted on specific disciplines (process, structural, mechanical, instrumentation) and available for project phases, not just permanent roles.") +
    p("Worth 20 minutes to see if there's a fit for what your team has coming up?", bottom=0)
),
"C": corps_block(
    p("Hi {first_name},") +
    p("I'll be straight with you: most {sector} companies we talk to say engineering talent is one of their top operational constraints right now.") +
    p("Trees Engineering is how some of the biggest names in the sector (Technip, Worley, Subsea7) close that gap quickly, without building out a full hire.") +
    p("Is that a conversation worth having at " + bold("{company}") + "?", bottom=0)
),
}

FINAL_CORPS = {
"A": corps_block(
    p("Hi {first_name},") +
    p("I've sent a couple of notes and haven't heard back, so I'll leave it here for now.") +
    p("If finding specialist engineering talent ever becomes a priority for the team at " + bold("{company}") + ", I'd genuinely love to show you what we've built at Trees.") +
    p("Take care, and best of luck with everything ahead.", bottom=0)
),
"B": corps_block(
    p("Hi {first_name},") +
    p("I've tried a couple of times without luck, so this will be my last message.") +
    p("If a project ever comes up where you need a specific engineer quickly and don't want to go through a 6-week agency process, Trees Engineering is worth knowing about.") +
    p("Wishing you a great rest of the year.", bottom=0)
),
"C": corps_block(
    p("Hi {first_name},") +
    p("I've reached out a few times and I respect that now's probably not the right moment.") +
    p("I'll leave it here. If engineering talent or project staffing ever becomes a real priority at " + bold("{company}") + ", you know where to find us.") +
    p("Good luck with everything on your end.", bottom=0)
),
}

FOLLOWUP_CTA = """<tr><td style="padding:24px 40px 32px;">
  <p style="margin:0 0 20px;font-size:14px;color:#444;line-height:1.7;">Alex is based in KL this month and has a few slots open for a quick intro call.</p>
  <table cellpadding="0" cellspacing="0" style="margin-bottom:10px;"><tr><td style="background-color:#0D1F35;padding:14px 28px;"><a href="https://calendar.app.google/bo4uFRj31jdTNdeW8" style="font-size:13px;font-weight:700;color:#ffffff;text-decoration:none;">Book a 20-min call with Alex</a></td></tr></table>
  <table cellpadding="0" cellspacing="0"><tr><td style="background-color:#25D366;padding:14px 28px;"><a href="https://wa.me/60174547710" style="font-size:13px;font-weight:700;color:#ffffff;text-decoration:none;">Message Alex on WhatsApp</a></td></tr></table>
</td></tr>"""

def build_html(version, sequence, first_name_str, company, sector):
    tpl_path = os.path.join(BASE_DIR, "templates", f"template_{version}.html")
    with open(tpl_path, "r", encoding="utf-8") as f:
        html = f.read()

    if sequence == "Follow-up 1":
        corps = FOLLOWUP1_CORPS[version]
    elif sequence == "Follow-up 2":
        corps = FOLLOWUP2_CORPS
    else:
        corps = FINAL_CORPS

    corps = corps.format(first_name=first_name_str, company=company, sector=sector)

    html = re.sub(r"<!-- CORPS.*?-->.*?<!-- WHAT WE DO -->",
                  corps + "\n        <!-- WHAT WE DO -->", html, flags=re.DOTALL)
    html = re.sub(r"<!-- CTA -->.*?<!-- SIGNATURES -->",
                  FOLLOWUP_CTA + "\n\n        <!-- SIGNATURES -->", html, flags=re.DOTALL)
    html = html.replace("{{FIRST_NAME}}", first_name_str).replace("{{COMPANY}}", company)
    return html

# ── EMAIL BUILDER (CID images) ────────────────────────────────────────────────

def build_mime(to_addr, subject, html_body):
    msg = MIMEMultipart("related")
    msg["From"]    = GMAIL
    msg["To"]      = to_addr
    msg["Subject"] = subject
    msg["Date"]    = email_lib.utils.formatdate(localtime=True)

    alt = MIMEMultipart("alternative")
    alt.attach(MIMEText("Please view this email in HTML mode.", "plain", "utf-8"))

    cid_html = html_body
    cid_parts = []
    def replace_data_uri(m):
        mime_type = m.group(1)
        b64_data  = m.group(2)
        cid = f"img_{len(cid_parts):02d}"
        cid_parts.append((cid, mime_type, b64_data))
        return f'src="cid:{cid}"'
    cid_html = re.sub(r'src="data:(image/[^;]+);base64,([A-Za-z0-9+/=]+)"', replace_data_uri, cid_html)

    alt.attach(MIMEText(cid_html, "html", "utf-8"))
    msg.attach(alt)

    for cid, mime_type, b64_data in cid_parts:
        img_data = base64.b64decode(b64_data)
        subtype = mime_type.split("/")[-1]
        part = MIMEImage(img_data, _subtype=subtype)
        part.add_header("Content-ID", f"<{cid}>")
        part.add_header("Content-Disposition", "inline")
        msg.attach(part)

    return msg

# ── LOAD CONTACTS ─────────────────────────────────────────────────────────────

def load_contacts():
    if not EXCEL_PATH or not os.path.exists(EXCEL_PATH):
        print(f"ERROR: Excel file not found: {EXCEL_PATH}")
        print("Set EXCEL_PATH in your .env file.")
        sys.exit(1)

    df = pd.read_excel(EXCEL_PATH, sheet_name="All Contacts", header=1, dtype=str)
    df.columns = df.columns.str.strip()
    results = []

    for idx, row in df.iterrows():
        status = row.get("Status", "")
        if not isinstance(status, str) or "Contacted" not in status: continue
        if "Whatsapp" in status: continue

        contact_date = parse_date(status)
        if not contact_date: continue

        days = (TODAY - contact_date).days
        sequence = get_sequence(days)
        if not sequence: continue

        title   = str(row.get("Title",   "") or "")
        company = str(row.get("Company", "") or "")
        version = get_version(title)
        sector  = get_sector(company)
        fname   = first_name(row.get("Full Name", ""))

        email_addr = str(row.get("Work Email", "") or "").strip()
        if not email_addr or email_addr == "nan":
            email_addr = str(row.get("Personal Email", "") or "").strip()
        if not email_addr or email_addr == "nan":
            email_addr = ""

        subject = SUBJECTS[sequence].replace("{sector}", sector)

        results.append({
            "row_idx":    idx + 2,
            "Full Name":  str(row.get("Full Name", "") or ""),
            "Company":    company,
            "Email":      email_addr,
            "Version":    version,
            "Sequence":   sequence,
            "Days":       days,
            "Subject":    subject,
            "first_name": fname,
            "sector":     sector,
        })

    return results, df

# ── UPDATE TRACKER ────────────────────────────────────────────────────────────

def update_tracker(contacts, label):
    wb = load_workbook(EXCEL_PATH)
    ws = wb["All Contacts"]
    headers = [ws.cell(row=2, column=c).value for c in range(1, 20)]
    try:
        fu_col = headers.index("Follow-up Status") + 1
    except ValueError:
        fu_col = len([h for h in headers if h]) + 1
        ws.cell(row=2, column=fu_col, value="Follow-up Status")

    for r in contacts:
        if r["Email"]:
            ws.cell(row=r["row_idx"], column=fu_col, value=label)
    wb.save(EXCEL_PATH)

# ── MODE: DRAFTS ──────────────────────────────────────────────────────────────

def mode_drafts(contacts):
    imap = imaplib.IMAP4_SSL("imap.gmail.com")
    imap.login(GMAIL, APP_PASSWORD)

    created = 0
    for r in contacts:
        if not r["Email"]:
            print(f"  SKIP (no email): {r['Full Name']}")
            continue
        html = build_html(r["Version"], r["Sequence"], r["first_name"], r["Company"], r["sector"])
        msg  = build_mime(r["Email"], r["Subject"], html)
        imap.append("[Gmail]/Drafts", "\\Draft",
                    imaplib.Time2Internaldate(datetime.now(timezone.utc)), msg.as_bytes())
        print(f"  DRAFT: {r['Full Name']} <{r['Email']}> [{r['Sequence']}]")
        created += 1

    imap.logout()
    update_tracker([r for r in contacts if r["Email"]],
                   f"Draft created ({TODAY.strftime('%d/%m/%Y')})")
    print(f"\n{created} brouillon(s) cree(s) dans Gmail.")

# ── MODE: SEND ────────────────────────────────────────────────────────────────

def mode_send(contacts):
    to_send = [r for r in contacts if r["Email"]]
    if not to_send:
        print("Aucun email a envoyer aujourd'hui.")
        return

    print(f"\n{len(to_send)} email(s) a envoyer. Debut dans 5 secondes...")
    time.sleep(5)

    smtp = smtplib.SMTP_SSL("smtp.gmail.com", 465)
    smtp.login(GMAIL, APP_PASSWORD)

    sent = 0
    for i, r in enumerate(to_send):
        html = build_html(r["Version"], r["Sequence"], r["first_name"], r["Company"], r["sector"])
        msg  = build_mime(r["Email"], r["Subject"], html)
        smtp.sendmail(GMAIL, r["Email"], msg.as_bytes())
        print(f"  SENT [{i+1}/{len(to_send)}]: {r['Full Name']} <{r['Email']}> [{r['Sequence']}]")
        sent += 1
        if i < len(to_send) - 1:
            print(f"  Attente 45s...")
            time.sleep(45)

    smtp.quit()
    update_tracker(to_send, f"Sent ({TODAY.strftime('%d/%m/%Y')})")
    print(f"\n{sent} email(s) envoye(s).")

# ── MODE: SCHEDULE ────────────────────────────────────────────────────────────

def mode_schedule():
    import subprocess
    script = os.path.abspath(__file__)
    python = sys.executable
    ps = f"""
$action  = New-ScheduledTaskAction -Execute '{python}' -Argument '"{script}" --mode drafts' -WorkingDirectory '{BASE_DIR}'
$trigger = New-ScheduledTaskTrigger -Daily -At '08:00AM'
$settings = New-ScheduledTaskSettingsSet -StartWhenAvailable -RunOnlyIfNetworkAvailable
Register-ScheduledTask -TaskName 'TreesEngineering_DailyDrafts' -Action $action -Trigger $trigger -Settings $settings -Force
Write-Host 'Tache planifiee creee : chaque matin a 08h00'
"""
    subprocess.run(["powershell", "-Command", ps], check=True)

# ── MAIN ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--mode", choices=["drafts","send","schedule"], default="drafts")
    args = parser.parse_args()

    print(f"=== Trees Engineering Outreach | {TODAY} | mode={args.mode} ===")

    if args.mode == "schedule":
        mode_schedule()
        sys.exit(0)

    contacts, _ = load_contacts()
    print(f"Contacts en fenetre de relance aujourd'hui : {len(contacts)}")

    if not contacts:
        print("Rien a faire aujourd'hui.")
        sys.exit(0)

    for r in contacts:
        flag = " [NO EMAIL]" if not r["Email"] else ""
        print(f"  [{r['Sequence']}] {r['Full Name']} | {r['Company']}{flag}")

    if args.mode == "drafts":
        mode_drafts(contacts)
    elif args.mode == "send":
        mode_send(contacts)
